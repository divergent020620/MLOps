import openpyxl
from copy import copy

SRC = r'C:\Users\27404\Desktop\BOS\cubeStudio\cube-studio-master\CubeStudio离线部署操作手册-SIT环境-v1.3.xlsx'
DST = r'C:\Users\27404\Desktop\BOS\cubeStudio\cube-studio-master\CubeStudio离线部署操作手册-SIT环境-v2.0.xlsx'

wb = openpyxl.load_workbook(SRC)

# ===== IP MAPPING =====
ip_map = {
    '10.240.9.13': '10.240.125.39',
    '10.240.9.14': '10.240.125.48',
    '10.240.9.15': '10.240.125.77',
    '10.240.9.16': '10.240.125.80',
    '10.240.9.49': '10.240.125.82',
}
worker3_ip = '10.240.125.83'
all_nodes_old = '10.240.9.13/14/15/16/49'
all_nodes_new = '10.240.125.39/48/77/80/82/83'

def replace_text(val):
    """Apply all text replacements to a cell value."""
    if val is None or not isinstance(val, str):
        return val
    s = val

    # ---- Step 1: Replace full shorthand patterns BEFORE individual IPs ----
    # "10.240.9.13/14/15/16/49" -> "10.240.125.39/48/77/80/82/83"
    s = s.replace('10.240.9.13/14/15/16/49', '10.240.125.39/48/77/80/82/83')
    # "10.240.9.14/15/16/49" -> "10.240.125.48/77/80/82/83"
    s = s.replace('10.240.9.14/15/16/49', '10.240.125.48/77/80/82/83')
    # "10.240.9.14/15" -> "10.240.125.48/77"
    s = s.replace('10.240.9.14/15', '10.240.125.48/77')
    # "10.240.9.16/49" -> "10.240.125.80/82/83"
    s = s.replace('10.240.9.16/49', '10.240.125.80/82/83')
    # /13/14/15/16/49 with leading IP already replaced, but other octet shorthands remain
    s = s.replace('10.240.125.39/14/15/16/49', '10.240.125.39/48/77/80/82/83')

    # ---- Step 2: Individual IP replacements ----
    for old_ip, new_ip in ip_map.items():
        s = s.replace(old_ip, new_ip)

    # ---- Step 3: Post-IP shorthand fixups (for patterns that got partially updated) ----
    # Fix /14 -> /48 (when standalone as shorthand, i.e., not part of a larger number)
    # These appear in patterns like "10.240.125.39/14" after main IP replacement
    # Use regex-like approach: replace "/14" in IP context -> "/48"
    import re
    # Fix shorthand octets that were NOT caught above
    s = re.sub(r'(?<=10\.240\.125\.39)/14\b', '/48', s)
    s = re.sub(r'(?<=10\.240\.125\.48)/15\b', '/77', s)
    s = re.sub(r'(?<=10\.240\.125\.77)/16\b', '/80', s)
    s = re.sub(r'(?<=10\.240\.125\.80)/49\b', '/82', s)
    # Also fix standalone /14/15/... sequences (shorthand without full IP prefix)
    # "/14" when in a chain like "/14/15/..." means 10.240.9.14 which maps to 10.240.125.48
    s = re.sub(r'\b/14\b', '/48', s)
    s = re.sub(r'\b/15\b', '/77', s)
    s = re.sub(r'\b/16\b', '/80', s)
    s = re.sub(r'\b/49\b', '/82', s)

    # ---- Step 4: OS / RPM / counts ----
    s = s.replace('CentOS Linux release 7.7.1908', 'Kylin Linux Advanced Server V10')
    s = s.replace('CentOS 7.7.1908', 'Kylin V10')
    s = s.replace('CentOS 7.x', 'Kylin V10')
    s = s.replace('centos', 'kilin')
    s = s.replace('3.10.0-1062.el7.x86_64', '4.19.90-52.40.v2207.ky10.x86_64')
    s = s.replace('rpm-new-centos', 'rpm-new-kilin')
    s = s.replace('3Master + 2Worker', '3Master + 3Worker')
    s = s.replace('3Master+2Worker', '3Master+3Worker')
    s = s.replace('5节点', '6节点')
    s = s.replace('5个节点', '6个节点')
    # Worker lists
    s = s.replace('（Worker-1、Worker-2）', '（Worker-1、Worker-2、Worker-3）')
    s = s.replace('Worker-1、Worker-2\n', 'Worker-1、Worker-2、Worker-3\n')
    # k8s-worker2 lines that should add worker3
    if '10.240.125.82 k8s-worker2' in s and '10.240.125.83' not in s:
        s = s.replace('10.240.125.82 k8s-worker2', '10.240.125.82 k8s-worker2\n10.240.125.83 k8s-worker3')

    # ---- Step 5: Fix ping lists (add missing worker3) ----
    if 'for ip in 10.240.125.48 10.240.125.77 10.240.125.80 10.240.125.82' in s:
        s = s.replace(
            'for ip in 10.240.125.48 10.240.125.77 10.240.125.80 10.240.125.82',
            'for ip in 10.240.125.48 10.240.125.77 10.240.125.80 10.240.125.82 10.240.125.83'
        )

    # ---- Step 6: Fix hostname setup (add worker3 hostname command) ----
    if '10.240.125.82 → hostnamectl set-hostname k8s-worker2' in s and \
       '10.240.125.83 → hostnamectl set-hostname k8s-worker3' not in s:
        s = s.replace(
            '10.240.125.82 → hostnamectl set-hostname k8s-worker2',
            '10.240.125.82 → hostnamectl set-hostname k8s-worker2\n10.240.125.83 → hostnamectl set-hostname k8s-worker3'
        )

    return s


# ===== PASS 1: Global text replacements =====
print("=== PASS 1: Global text replacements ===")
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                new_val = replace_text(cell.value)
                if new_val != cell.value:
                    cell.value = new_val
    print(f"  Done: {sheet_name}")

# ===== PASS 2: Add worker3 to 服务器与软件版本 =====
print("\n=== PASS 2: Adding worker3 to server sheet ===")
ws_ver = wb['服务器与软件版本']
for row in range(1, ws_ver.max_row + 1):
    cell_val = ws_ver.cell(row=row, column=1).value
    if cell_val and '10.240.125.82' in str(cell_val):
        ws_ver.insert_rows(row + 1, 1)
        for col in range(1, 7):
            src_cell = ws_ver.cell(row=row, column=col)
            dst_cell = ws_ver.cell(row=row + 1, column=col)
            if src_cell.has_style:
                dst_cell.font = copy(src_cell.font)
                dst_cell.border = copy(src_cell.border)
                dst_cell.fill = copy(src_cell.fill)
                dst_cell.number_format = src_cell.number_format
                dst_cell.protection = copy(src_cell.protection)
                dst_cell.alignment = copy(src_cell.alignment)
        ws_ver.cell(row=row + 1, column=1).value = '10.240.125.83'
        ws_ver.cell(row=row + 1, column=2).value = 'k8s-worker3'
        ws_ver.cell(row=row + 1, column=3).value = 'Worker'
        ws_ver.cell(row=row + 1, column=4).value = '40C/256G'
        ws_ver.cell(row=row + 1, column=5).value = 'Kylin V10'
        print(f"  Added worker3 at row {row + 1}")
        break

# ===== PASS 3: NFS reorder in K8s集群部署 =====
print("\n=== PASS 3: NFS reorder ===")
ws_k8s = wb['K8s集群部署']

# Update overview
ws_k8s.cell(row=4, column=4).value = '前置检查→环境准备→NFS→Containerd→Harbor(镜像仓库)→K8s→Flannel'

# K8s sheet rows:
# Row 8: headers
# Row 9-14: 准备阶段 steps 1-6
# Row 15-31: 部署阶段 steps 1-17 (17 rows)
# Row 32: 验证阶段
# Row 33-34: 回退阶段

# Read all deployment step rows (15-31) into memory
deploy_rows = []
for row_num in range(15, 32):
    row_data = {}
    for col in range(1, 13):  # columns A-L
        cell = ws_k8s.cell(row=row_num, column=col)
        row_data[col] = {
            'value': cell.value,
            'font': copy(cell.font),
            'border': copy(cell.border),
            'fill': copy(cell.fill),
            'number_format': cell.number_format,
            'protection': copy(cell.protection),
            'alignment': copy(cell.alignment),
        }
    deploy_rows.append(row_data)

# Original order: [0]=Containerd(step1), [1]=Config(step2), ..., [16]=NFS(step17)
# New order: NFS first, then Containerd, then rest
# NFS was at index 16, move to index 0
nfs_row = deploy_rows.pop(16)  # Remove NFS (last item)
deploy_rows.insert(0, nfs_row)  # Insert NFS at beginning

# Write back with updated step numbers
# Column A (row 15 only) is the merged "部署阶段" label; rows 16-31 col A are MergedCells (read-only)
for i, row_data in enumerate(deploy_rows):
    row_num = 15 + i
    new_step_num = i + 1
    cols_to_write = list(range(2, 13))  # B-L
    if row_num == 15:
        cols_to_write = [1] + cols_to_write  # Include col A only for row 15 (top-left of merged range)
    for col in cols_to_write:
        cell = ws_k8s.cell(row=row_num, column=col)
        cell.value = row_data[col]['value']
        cell.font = row_data[col]['font']
        cell.border = row_data[col]['border']
        cell.fill = row_data[col]['fill']
        cell.number_format = row_data[col]['number_format']
        cell.protection = row_data[col]['protection']
        cell.alignment = row_data[col]['alignment']
    # Update step number in column B
    ws_k8s.cell(row=row_num, column=2).value = new_step_num

print("  NFS moved to step 1, Containerd shifted to step 2")

# ---- 3.1: Add worker3 to Worker join step (now row 29 after reorder) ----
ws_k8s.cell(row=29, column=4).value = '10.240.125.80\n(k8s-worker1)\n10.240.125.82\n(k8s-worker2)\n10.240.125.83\n(k8s-worker3)'
print("  Worker3 added to join step")

# ===== PASS 4: Update 物料清单 node counts =====
print("\n=== PASS 4: Update material list ===")
ws_mat = wb['物料清单']
for row in range(1, ws_mat.max_row + 1):
    cell_c = ws_mat.cell(row=row, column=3)
    if cell_c.value is not None:
        # Replace "5" with "6" for node count column
        if isinstance(cell_c.value, str) and cell_c.value.strip() == '5':
            cell_c.value = '6'
        elif isinstance(cell_c.value, (int, float)) and cell_c.value == 5:
            cell_c.value = 6
    # Also update descriptions that mention node count
    cell_b = ws_mat.cell(row=row, column=2)
    if cell_b.value and isinstance(cell_b.value, str):
        cell_b.value = cell_b.value.replace('5个', '6个')

# ===== PASS 5: CubeStudio sheet - update worker3 references =====
print("\n=== PASS 5: CubeStudio sheet updates ===")
ws_cube = wb['CubeStudio部署']
for row in range(1, ws_cube.max_row + 1):
    for col in range(1, ws_cube.max_column + 1):
        cell = ws_cube.cell(row=row, column=col)
        if cell.value and isinstance(cell.value, str):
            s = cell.value
            if 'k8s-worker2' in s and 'k8s-worker3' not in s:
                s = s.replace('k8s-worker2', 'k8s-worker2\n10.240.125.83 k8s-worker3')
                cell.value = s

# ===== PASS 6: Update 附录 port/directory listings =====
print("\n=== PASS 6: Appendix updates ===")
ws_app = wb['附录']
for row in range(1, ws_app.max_row + 1):
    for col in range(1, ws_app.max_column + 1):
        cell = ws_app.cell(row=row, column=col)
        if cell.value and isinstance(cell.value, str):
            s = cell.value
            # Update IP references in directory structure
            if '10.240.125.39' in s or 'k8s-worker' in s:
                if 'k8s-worker2' in s and 'k8s-worker3' not in s:
                    s = s.replace('k8s-worker2', 'k8s-worker2、k8s-worker3')
                    cell.value = s

# ===== Save =====
wb.save(DST)
print(f"\n=== Saved to: {DST} ===")
print("Done!")
