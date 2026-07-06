import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==================== STYLES ====================
FONT_TITLE = Font(name='Arial', size=14, bold=True, color='FF222222')
FONT_SECTION = Font(name='Arial', size=11, bold=True, color='FF222222')
FONT_HEADER = Font(name='Arial', size=11, bold=True, color='FFFFFFFF')
FONT_NORMAL = Font(name='Arial', size=11, bold=False, color='FF222222')
FONT_INDEX = Font(name='Arial', size=11, bold=True, color='FF222222')
FONT_HOST = Font(name='Arial', size=10, bold=True, color='FF222222')
FONT_MONO = Font(name='Consolas', size=10, color='FF222222')

FILL_TITLE = PatternFill(patternType='solid', fgColor='FF4472C4')
FILL_HEADER = PatternFill(patternType='solid', fgColor='FF5B9BD5')
FILL_SECTION = PatternFill(patternType='solid', fgColor='FFD9E2F3')
FILL_INFO_LABEL = PatternFill(patternType='solid', fgColor='FFD9E2F3')
FILL_PHASE = PatternFill(patternType='solid', fgColor='FFE2EFDA')
FILL_ALT_ROW = PatternFill(patternType='solid', fgColor='FFF2F2F2')

ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrapText=True)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrapText=True)
ALIGN_LEFT_TOP = Alignment(horizontal='left', vertical='top', wrapText=True)

THIN_BORDER = Border(
    left=Side(style='thin', color='FFBFBFBF'),
    right=Side(style='thin', color='FFBFBFBF'),
    top=Side(style='thin', color='FFBFBFBF'),
    bottom=Side(style='thin', color='FFBFBFBF'),
)

def ac(ws, row, col, value, font=FONT_NORMAL, fill=None, alignment=ALIGN_LEFT):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font; cell.border = THIN_BORDER
    if fill: cell.fill = fill
    cell.alignment = alignment
    return cell

def set_col_widths(ws, widths):
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i+1)].width = w

def set_row_h(ws, row, height):
    ws.row_dimensions[row].height = height

def merge_title(ws, row, cols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    ac(ws, row, 1, text, font=FONT_TITLE, fill=FILL_TITLE, alignment=ALIGN_CENTER)
    set_row_h(ws, row, 30)

def write_header_row(ws, row, headers):
    for i, h in enumerate(headers):
        ac(ws, row, i+1, h, font=FONT_HEADER, fill=FILL_HEADER, alignment=ALIGN_CENTER)
    set_row_h(ws, row, 25)

def freeze_and_print(ws, freeze_cell):
    ws.freeze_panes = freeze_cell
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0

def write_phase_steps(ws, start_row, rows_data):
    r = start_row
    for ridx, (phase, seq, purpose, hostname, user, commands, verify) in enumerate(rows_data):
        row_num = r + ridx
        ac(ws, row_num, 1, phase, font=FONT_NORMAL, fill=FILL_PHASE, alignment=ALIGN_CENTER)
        ac(ws, row_num, 2, seq, font=FONT_INDEX, alignment=ALIGN_CENTER)
        ac(ws, row_num, 3, purpose, font=FONT_NORMAL, alignment=ALIGN_LEFT_TOP)
        ac(ws, row_num, 4, hostname, font=FONT_HOST, alignment=ALIGN_LEFT_TOP)
        ac(ws, row_num, 5, user, font=FONT_NORMAL, alignment=ALIGN_CENTER)
        ac(ws, row_num, 6, commands, font=FONT_NORMAL, alignment=ALIGN_LEFT_TOP)
        ac(ws, row_num, 7, '', font=FONT_NORMAL, alignment=ALIGN_CENTER)
        ac(ws, row_num, 8, '', font=FONT_NORMAL, alignment=ALIGN_CENTER)
        ac(ws, row_num, 9, '', font=FONT_NORMAL, alignment=ALIGN_CENTER)
        ac(ws, row_num, 10, verify, font=FONT_NORMAL, alignment=ALIGN_LEFT_TOP)
        ac(ws, row_num, 11, '', font=FONT_NORMAL, alignment=ALIGN_CENTER)
        max_lines = max(len(commands.split('\n')), len(verify.split('\n')), len(hostname.split('\n')))
        set_row_h(ws, row_num, max(25, min(400, max_lines * 15)))

    phase_groups = {}
    for ridx, (phase, _, _, _, _, _, _) in enumerate(rows_data):
        if phase not in phase_groups:
            phase_groups[phase] = [ridx, ridx]
        else:
            phase_groups[phase][1] = ridx
    for phase, (start, end) in phase_groups.items():
        if start != end:
            ws.merge_cells(start_row=r+start, start_column=1, end_row=r+end, end_column=1)
    return r + len(rows_data)


# ==================== CONSTANTS ====================
ALL = '所有节点\n10.240.9.13/14/15/16/49'
M1 = '10.240.9.13\n(k8s-master1)'
M23 = '10.240.9.14\n(k8s-master2)\n10.240.9.15\n(k8s-master3)'
W = '10.240.9.16\n(k8s-worker1)\n10.240.9.49\n(k8s-worker2)'
M23W = '10.240.9.14\n(k8s-master2)\n10.240.9.15\n(k8s-master3)\n10.240.9.16\n(k8s-worker1)\n10.240.9.49\n(k8s-worker2)'

MAIN_COLS = 11
HDR = ['阶段', '序号', '操作内容/目的', '目标主机名', '执行用户名',
       '详细步骤描述\n（细化到命令行）', '开始时间', '完成时间', '操作人',
       '验证方法/步骤简要描述', '复核人']

# ==================== WORKBOOK ====================
wb = openpyxl.Workbook()

# ========================================
# Sheet 1: K8s集群部署（含所有前置步骤 + K8s + Harbor + Flannel + NFS）
# ========================================
ws1 = wb.active
ws1.title = 'K8s集群部署'
set_col_widths(ws1, [10, 7, 20, 22, 10, 55, 14, 14, 10, 40, 10])

r = 1
merge_title(ws1, r, MAIN_COLS, 'Kubernetes集群部署 — v1.28.2（3Master + 2Worker + Harbor）')
r += 1
ws1.merge_cells(start_row=r, start_column=1, end_row=r+2, end_column=1)
ac(ws1, r, 1, '部署概况', font=FONT_SECTION, fill=FILL_SECTION, alignment=ALIGN_CENTER)
for i, (idx, label, content) in enumerate([
    (1, 'K8s版本', 'v1.28.2，kubeadm部署，3Master+2Worker高可用架构，Harbor作为集群内镜像仓库'),
    (2, '容器运行时/网络', 'Containerd 1.7.29 / Flannel v0.28.4'),
    (3, '操作内容', '前置检查→环境准备(主机名/hosts/防火墙/SELinux/swap/内核)→Containerd→Harbor→K8s→Flannel→NFS'),
]):
    ac(ws1, r+i, 2, idx, font=FONT_INDEX, fill=FILL_SECTION, alignment=ALIGN_CENTER)
    ac(ws1, r+i, 3, label, font=FONT_NORMAL, fill=FILL_INFO_LABEL)
    ws1.merge_cells(start_row=r+i, start_column=4, end_row=r+i, end_column=MAIN_COLS)
    ac(ws1, r+i, 4, content, font=FONT_NORMAL)
r += 3

ws1.merge_cells(start_row=r, start_column=1, end_row=r+2, end_column=1)
ac(ws1, r, 1, '回退决策条件', font=FONT_SECTION, fill=FILL_SECTION, alignment=ALIGN_CENTER)
for i, (idx, content) in enumerate([
    (1, 'K8s节点异常超过2个无法恢复时，执行kubeadm reset回退'),
    (2, '初始化失败时，kubeadm reset -f后重新初始化'),
    (3, '部署时间超过预期窗口时，由负责人决策是否回退'),
]):
    ac(ws1, r+i, 2, idx, font=FONT_INDEX, fill=FILL_SECTION, alignment=ALIGN_CENTER)
    ws1.merge_cells(start_row=r+i, start_column=3, end_row=r+i, end_column=MAIN_COLS)
    ac(ws1, r+i, 3, content, font=FONT_NORMAL)
r += 3

write_header_row(ws1, r, HDR)
r += 1

k8s_rows = [
    # ===== 准备阶段 =====
    ('准备阶段', 1, '前置检查', ALL, 'root',
     '''# 系统版本
cat /etc/redhat-release  # CentOS Linux release 7.7.1908
uname -r                 # 3.10.0-1062.el7.x86_64
# 资源
nproc                    # 40核
free -g | grep Mem       # 256GB
df -h /                  # > 100GB
# 网络
for ip in 10.240.9.14 10.240.9.15 10.240.9.16 10.240.9.49; do ping -c 2 $ip; done
# 端口检查（K8s关键端口，应无输出）
netstat -tunlp | grep -E ':(6443|2379|2380|10250|10251|10252|10255)'
# 旧环境检查
ps aux | grep -E 'kube|etcd' | grep -v grep
ls -la /etc/kubernetes/ 2>/dev/null
systemctl status firewalld --no-pager
systemctl status iptables --no-pager
getenforce''',
     '''系统 CentOS 7.7.1908；所有IP可达
关键端口未占用；无旧K8s残留'''),

    ('准备阶段', 2, '设置主机名', ALL, 'root',
     '''# 各服务器分别执行
10.240.9.13 → hostnamectl set-hostname k8s-master1
10.240.9.14 → hostnamectl set-hostname k8s-master2
10.240.9.15 → hostnamectl set-hostname k8s-master3
10.240.9.16 → hostnamectl set-hostname k8s-worker1
10.240.9.49 → hostnamectl set-hostname k8s-worker2''',
     'hostname  # 确认主机名已更改'),

    ('准备阶段', 3, '配置hosts文件', ALL, 'root',
     '''cp /etc/hosts /etc/hosts.bak
cat > /etc/hosts << 'EOF'
127.0.0.1   localhost localhost.localdomain
::1         localhost localhost.localdomain

10.240.9.13 k8s-master1
10.240.9.14 k8s-master2
10.240.9.15 k8s-master3
10.240.9.16 k8s-worker1
10.240.9.49 k8s-worker2
EOF''',
     'cat /etc/hosts'),

    ('准备阶段', 4, '关闭防火墙与SELinux', ALL, 'root',
     '''systemctl stop firewalld && systemctl disable firewalld
systemctl stop iptables 2>/dev/null
setenforce 0
sed -i 's/^SELINUX=enforcing/SELINUX=disabled/' /etc/selinux/config''',
     '''getenforce              # Disabled
systemctl status firewalld  # inactive'''),

    ('准备阶段', 5, '关闭swap', ALL, 'root',
     '''swapoff -a
cp /etc/fstab /etc/fstab.bak
sed -i '/swap/s/^/#/' /etc/fstab''',
     'free -h | grep Swap  # 应为0'),

    ('准备阶段', 6, '配置内核参数\n及内核模块', ALL, 'root',
     '''yum install -y yum-utils device-mapper-persistent-data lvm2
yum install -y iptables container-selinux iptables-services

# 写入rc.local（防火墙禁用+内核模块加载）
(
cat << EOF
systemctl stop firewalld
systemctl disable firewalld
systemctl stop iptables
systemctl disable iptables
systemctl stop ip6tables
systemctl disable ip6tables
systemctl stop nftables
systemctl disable nftables

modprobe br_netfilter
modprobe ip_tables
modprobe iptable_nat
modprobe iptable_filter
modprobe iptable_mangle
modprobe iptable_mangle
modprobe ip6_tables
modprobe ip6table_nat
modprobe ip6table_filter
modprobe ip6table_mangle
modprobe ip6table_mangle
EOF
) >> /etc/rc.d/rc.local
chmod +x /etc/rc.d/rc.local
sh /etc/rc.d/rc.local

lsmod
sudo echo 'ip_tables' >> /etc/modules

echo "net.bridge.bridge-nf-call-ip6tables = 1" >> /etc/sysctl.conf
echo "net.bridge.bridge-nf-call-iptables=1" >> /etc/sysctl.conf
echo "net.ipv4.ip_forward=1" >> /etc/sysctl.conf
echo "1" >/proc/sys/net/bridge/bridge-nf-call-iptables
sysctl -p

systemctl restart docker
reboot''',
     '''重启后验证：
getenforce                    # Disabled
systemctl status firewalld    # inactive
lsmod | grep br_netfilter     # 已加载
sysctl net.ipv4.ip_forward    # 1'''),

    # ===== 部署阶段：容器运行时 =====
    ('部署阶段', 1, '安装Containerd\n+ runc + CNI插件', ALL, 'root',
     '''# 解压nerdctl-full（含containerd、buildkit等）
tar -xzf /data/offline/nerdctl-full-1.7.6-linux-amd64.tar.gz -C /usr/local/

# 安装runc
install -m 755 /data/offline/containerd/runc.amd64 /usr/local/sbin/runc

# 安装CNI插件
mkdir -p /opt/cni/bin
tar -xzf /data/offline/cni/cni-plugins-linux-amd64-v1.1.1.tgz -C /opt/cni/bin/

# 复制systemd服务文件并修改buildkit
cp /usr/local/lib/systemd/system/*.service /etc/systemd/system/
# 修改 /etc/systemd/system/buildkit.service 的 ExecStart 为：
# ExecStart=/usr/local/bin/buildkitd --oci-worker=false --containerd-worker=true''',
     '''ls /usr/local/bin/containerd
ls /usr/local/sbin/runc
ls /opt/cni/bin/'''),

    ('部署阶段', 2, '配置Containerd', ALL, 'root',
     '''mkdir -p /etc/containerd
containerd config default > /etc/containerd/config.toml

# 修改 /etc/containerd/config.toml 关键改动：
# 1. sandbox_image = "registry.cn-hangzhou.aliyuncs.com/google_containers/pause:3.8"
# 2. SystemdCgroup = true
# 3. config_path = "/etc/containerd/certs.d"

# 创建docker.io镜像源hosts
mkdir -p /etc/containerd/certs.d/docker.io
tee /etc/containerd/certs.d/docker.io/hosts.toml << 'EOF'
server = "https://docker.io"
[host."https://docker.1panel.live"]
  capabilities = ["pull", "resolve"]
[host."https://hub.rat.dev/"]
  capabilities = ["pull", "resolve"]
[host."https://docker.chenby.cn"]
  capabilities = ["pull", "resolve"]
[host."https://docker.m.daocloud.io"]
  capabilities = ["pull", "resolve"]
EOF''',
     '''cat /etc/containerd/config.toml | grep -E "sandbox_image|SystemdCgroup|config_path"
cat /etc/containerd/certs.d/docker.io/hosts.toml'''),

    ('部署阶段', 3, '配置Harbor镜像源\n（Harbor部署后再执行）', ALL, 'root',
     '''mkdir -p /etc/containerd/certs.d/10.240.9.13
cat > /etc/containerd/certs.d/10.240.9.13/hosts.toml << 'EOF'
server = "https://10.240.9.13"
[host."https://10.240.9.13"]
  capabilities = ["pull", "resolve", "push"]
  skip_verify = true
EOF''',
     'cat /etc/containerd/certs.d/10.240.9.13/hosts.toml'),

    ('部署阶段', 4, '启动Containerd', ALL, 'root',
     '''systemctl daemon-reload
systemctl enable containerd buildkit --now
systemctl status containerd''',
     '''systemctl status containerd  # active (running)
ctr version'''),

    ('部署阶段', 5, '导入容器镜像\n（共54个.tar，仅Master-1）', M1, 'root',
     '''# Master-1导入镜像到containerd（用于kubeadm init + Harbor docker load推送）
cd /data/offline/images
for f in *.tar; do
  echo "导入: $f"
  ctr -n k8s.io images import "$f"
done
# 其他节点不导入，后续通过Harbor拉取镜像''',
     'ctr -n k8s.io images ls | wc -l  # 54'),

    # ===== 部署阶段：Harbor =====
    ('部署阶段', 6, '安装Docker\n（Harbor依赖）', M1, 'root',
     '''# 离线安装docker（rpm在rpm-new-centos目录中）
cd /data/offline/rpm-new-centos
rpm -ivh docker-ce-*.rpm docker-ce-cli-*.rpm containerd.io-*.rpm docker-compose-plugin-*.rpm --nodeps 2>/dev/null
# 如果离线包不全，有yum源时：
# yum install -y docker-ce docker-ce-cli containerd.io docker-compose-plugin
systemctl enable docker --now''',
     '''docker version
docker ps'''),

    ('部署阶段', 7, '安装Harbor镜像仓库', M1, 'root',
     '''cd /data/offline/harbor
tar -xzf harbor-offline-installer-v2.11.1.tgz -C /opt/
cd /opt/harbor
cp harbor.yml.tmpl harbor.yml
# 修改harbor.yml关键项：
# hostname: 10.240.9.13
# http.port: 80
# https注释掉（内网使用http）
# harbor_admin_password: Harbor12345
./install.sh''',
     '''docker ps | grep harbor
curl http://10.240.9.13/api/v2.0/health
默认账号 admin / Harbor12345'''),

    ('部署阶段', 8, '推送镜像到Harbor', M1, 'root',
     '''docker login 10.240.9.13 -u admin -p Harbor12345

curl -u admin:Harbor12345 -X POST http://10.240.9.13/api/v2.0/projects \\
  -H "Content-Type: application/json" \\
  -d '{"project_name":"cube-studio","public":true}'

# docker load所有镜像
for f in /data/offline/images/*.tar; do
  docker load -i "$f"
done

# 将所有镜像重新tag并push到Harbor
for img in $(docker images --format '{{.Repository}}:{{.Tag}}' | grep -v '<none>'); do
  new_tag="10.240.9.13/cube-studio/$img"
  docker tag "$img" "$new_tag"
  docker push "$new_tag"
done''',
     '浏览器访问Harbor确认cube-studio项目中有镜像'),

    # ===== 部署阶段：其他节点配置Harbor镜像源 =====
    ('部署阶段', 9, '配置Harbor镜像源\n（Master-2/3、Worker节点）', M23W, 'root',
     '''# 配置containerd Harbor镜像源（已在步骤3创建，确认存在）
cat /etc/containerd/certs.d/10.240.9.13/hosts.toml

# 配置K8s源registries通过Harbor拉取镜像
# 为 registry.cn-hangzhou.aliyuncs.com 创建Harbor mirror
mkdir -p /etc/containerd/certs.d/registry.cn-hangzhou.aliyuncs.com
cat > /etc/containerd/certs.d/registry.cn-hangzhou.aliyuncs.com/hosts.toml << 'EOF'
server = "https://registry.cn-hangzhou.aliyuncs.com"
[host."https://10.240.9.13"]
  capabilities = ["pull", "resolve"]
  skip_verify = true
EOF

# 为 ccr.ccs.tencentyun.com 创建Harbor mirror
mkdir -p /etc/containerd/certs.d/ccr.ccs.tencentyun.com
cat > /etc/containerd/certs.d/ccr.ccs.tencentyun.com/hosts.toml << 'EOF'
server = "https://ccr.ccs.tencentyun.com"
[host."https://10.240.9.13"]
  capabilities = ["pull", "resolve"]
  skip_verify = true
EOF

# 为 ghcr.io 创建Harbor mirror
mkdir -p /etc/containerd/certs.d/ghcr.io
cat > /etc/containerd/certs.d/ghcr.io/hosts.toml << 'EOF'
server = "https://ghcr.io"
[host."https://10.240.9.13"]
  capabilities = ["pull", "resolve"]
  skip_verify = true
EOF

# 为 docker.io 创建Harbor mirror（已有步骤2的公共源，增加Harbor）
cat >> /etc/containerd/certs.d/docker.io/hosts.toml << 'EOF'
[host."https://10.240.9.13"]
  capabilities = ["pull", "resolve"]
  skip_verify = true
EOF

systemctl restart containerd''',
     '''ls /etc/containerd/certs.d/registry.cn-hangzhou.aliyuncs.com/hosts.toml
ls /etc/containerd/certs.d/ccr.ccs.tencentyun.com/hosts.toml
ls /etc/containerd/certs.d/ghcr.io/hosts.toml
systemctl status containerd'''),

    # ===== 部署阶段：K8s =====
    ('部署阶段', 10, '安装K8s二进制文件', ALL, 'root',
     '''cd /data/offline/k8s-bin
install -m 755 kubeadm /usr/bin/kubeadm
install -m 755 kubelet /usr/bin/kubelet
install -m 755 kubectl /usr/bin/kubectl''',
     '''kubeadm version
kubelet --version
kubectl version --client'''),

    ('部署阶段', 11, '配置kubelet服务', ALL, 'root',
     '''cat > /etc/systemd/system/kubelet.service << 'EOF'
[Unit]
Description=kubelet: The Kubernetes Node Agent
Wants=network-online.target
After=network-online.target
[Service]
ExecStart=/usr/bin/kubelet
Restart=always
StartLimitInterval=0
RestartSec=10
[Install]
WantedBy=multi-user.target
EOF

mkdir -p /etc/systemd/system/kubelet.service.d
cat > /etc/systemd/system/kubelet.service.d/10-kubeadm.conf << 'EOF'
[Service]
Environment="KUBELET_KUBECONFIG_ARGS=--bootstrap-kubeconfig=/etc/kubernetes/bootstrap-kubelet.conf --kubeconfig=/etc/kubernetes/kubelet.conf"
Environment="KUBELET_CONFIG_ARGS=--config=/var/lib/kubelet/config.yaml"
EnvironmentFile=-/var/lib/kubelet/kubeadm-flags.env
EnvironmentFile=-/etc/default/kubelet
ExecStart=
ExecStart=/usr/bin/kubelet $KUBELET_KUBECONFIG_ARGS $KUBELET_CONFIG_ARGS $KUBELET_KUBEADM_ARGS $KUBELET_EXTRA_ARGS
EOF

systemctl daemon-reload
systemctl enable kubelet''',
     'systemctl status kubelet  # loaded状态'),

    ('部署阶段', 12, '初始化第一个Master\n（k8s-master1）', M1, 'root',
     '''cat > /root/kubeadm-init.yaml << 'EOF'
apiVersion: kubeadm.k8s.io/v1beta3
kind: InitConfiguration
localAPIEndpoint:
  advertiseAddress: 10.240.9.13
  bindPort: 6443
nodeRegistration:
  criSocket: unix:///run/containerd/containerd.sock
  name: k8s-master1
---
apiVersion: kubeadm.k8s.io/v1beta3
kind: ClusterConfiguration
kubernetesVersion: v1.28.2
controlPlaneEndpoint: "10.240.9.13:6443"
networking:
  serviceSubnet: "10.96.0.0/12"
  podSubnet: "10.244.0.0/16"
imageRepository: registry.cn-hangzhou.aliyuncs.com/google_containers
apiServer:
  certSANs:
  - "10.240.9.13"
  - "10.240.9.14"
  - "10.240.9.15"
  - "k8s-master1"
  - "k8s-master2"
  - "k8s-master3"
etcd:
  local:
    dataDir: /var/lib/etcd
---
apiVersion: kubelet.config.k8s.io/v1beta1
kind: KubeletConfiguration
cgroupDriver: systemd
EOF

kubeadm init --config=/root/kubeadm-init.yaml --upload-certs''',
     '''初始化成功输出 "Your Kubernetes control-plane has initialized successfully!"
保存输出的 kubeadm join 命令（Master加入命令和Worker加入命令）'''),

    ('部署阶段', 13, '配置kubectl', M1, 'root',
     '''mkdir -p /root/.kube
cp /etc/kubernetes/admin.conf /root/.kube/config
kubectl get nodes''',
     'kubectl get nodes  # 应显示k8s-master1节点'),

    ('部署阶段', 14, '加入其他Master节点\n（Master-2、Master-3）', M23, 'root',
     '''# 使用Master-1 init输出的Master加入命令：
kubeadm join 10.240.9.13:6443 --token <token> \\
    --discovery-token-ca-cert-hash sha256:<hash> \\
    --control-plane --certificate-key <cert-key>''',
     '''kubectl get nodes  # 应有3个Master节点
状态NotReady，等待CNI'''),

    ('部署阶段', 15, '加入Worker节点\n（Worker-1、Worker-2）', W, 'root',
     '''# 使用Master-1 init输出的Worker加入命令：
kubeadm join 10.240.9.13:6443 --token <token> \\
    --discovery-token-ca-cert-hash sha256:<hash>
# Token过期时在Master-1执行：kubeadm token create --print-join-command''',
     'kubectl get nodes  # 应有5个节点'),

    ('部署阶段', 16, '部署CNI网络插件\n（Flannel）', M1, 'root',
     '''kubectl apply -f /data/offline/kube-flannel.yml
kubectl get nodes -w    # 等待所有节点变为Ready''',
     'kubectl get nodes  # 所有节点Ready'),

    ('部署阶段', 17, '创建命名空间\n与持久化存储目录', ALL, 'root',
     '''# Master-1执行
kubectl create namespace cube-studio
# 所有节点
mkdir -p /data/k8s-volumes/{mysql,redis,cubestudio}
chmod -R 777 /data/k8s-volumes''',
     '''kubectl get ns cube-studio
ls -la /data/k8s-volumes/'''),

    ('部署阶段', 18, '配置NFS共享存储', ALL, 'root',
     '''# NFS Server（Master-1）
cd /data/offline && tar -zxvf nfsrpm.tar.gz
cd nfs && rpm -ivh *.rpm --force --nodeps
mkdir -p /data/nfs
echo "/data/nfs/ *(rw,no_root_squash,async)" >> /etc/exports
exportfs -arv
systemctl enable rpcbind nfs-server --now
mkdir -p /data/nfs/k8s && ln -s /data/nfs/k8s /data/

# NFS Client（Master-2/3、Worker节点）
cd /data/offline && tar -zxvf nfsrpm.tar.gz
cd nfs && rpm -ivh *.rpm --force --nodeps
showmount -e 10.240.9.13
mkdir -p /data/nfs
echo "10.240.9.13:/data/nfs  /data/nfs   nfs   defaults  0  0" >> /etc/fstab
mount -a
mkdir -p /data/nfs/k8s && ln -s /data/nfs/k8s /data/''',
     '''# Server端：showmount -e localhost
# Client端：showmount -e 10.240.9.13
df -h | grep nfs'''),

    # ===== 验证阶段 =====
    ('验证阶段', 1, '集群状态验证', M1, 'root',
     '''kubectl get nodes                    # 5节点全部Ready
kubectl get pods -n kube-system      # 全部Running
kubectl get pods -n kube-flannel     # 全部Running
kubectl cluster-info''',
     '5节点Ready；kube-system和kube-flannel Pod全部Running'),

    # ===== 回退阶段 =====
    ('回退阶段', 1, '清理K8s环境', '对应节点', 'root',
     '''kubeadm reset -f
rm -rf /etc/kubernetes/ /var/lib/etcd/ /var/lib/kubelet/ /var/lib/cni/ /etc/cni/net.d/ ~/.kube/
iptables -F && iptables -t nat -F && iptables -t mangle -F && iptables -X''',
     'ps aux | grep -E "kube|etcd" | grep -v grep  # 无残留'),

    ('回退阶段', 2, '恢复系统配置', ALL, 'root',
     '''cp /etc/hosts.bak /etc/hosts
cp /etc/fstab.bak /etc/fstab
rm -f /etc/sysctl.d/k8s.conf''',
     'cat /etc/hosts  # 已恢复原始内容'),
]

r = write_phase_steps(ws1, r, k8s_rows)
freeze_and_print(ws1, 'A9')

# ========================================
# Sheet 2: CubeStudio部署（仅CubeStudio软件）
# ========================================
ws2 = wb.create_sheet('CubeStudio部署')
set_col_widths(ws2, [10, 7, 20, 22, 10, 55, 14, 14, 10, 40, 10])

r = 1
merge_title(ws2, r, MAIN_COLS, 'CubeStudio平台部署')
r += 1
ws2.merge_cells(start_row=r, start_column=1, end_row=r+2, end_column=1)
ac(ws2, r, 1, '部署概况', font=FONT_SECTION, fill=FILL_SECTION, alignment=ALIGN_CENTER)
for i, (idx, label, content) in enumerate([
    (1, '背景及目的', '在SIT测试环境的K8s集群上部署CubeStudio云原生机器学习平台'),
    (2, '部署前提', 'K8s集群已部署完成（3Master+2Worker+Harbor，参见K8s集群部署Sheet）'),
    (3, '本次操作内容', '解压源码→配置调整→执行部署→验证→回退'),
]):
    ac(ws2, r+i, 2, idx, font=FONT_INDEX, fill=FILL_SECTION, alignment=ALIGN_CENTER)
    ac(ws2, r+i, 3, label, font=FONT_NORMAL, fill=FILL_INFO_LABEL)
    ws2.merge_cells(start_row=r+i, start_column=4, end_row=r+i, end_column=MAIN_COLS)
    ac(ws2, r+i, 4, content, font=FONT_NORMAL)
r += 3

ws2.merge_cells(start_row=r, start_column=1, end_row=r+1, end_column=1)
ac(ws2, r, 1, '回退决策条件', font=FONT_SECTION, fill=FILL_SECTION, alignment=ALIGN_CENTER)
for i, (idx, content) in enumerate([
    (1, 'CubeStudio平台核心服务持续无法启动时，删除namespace回退'),
    (2, 'K8s集群异常时，参考K8s集群部署Sheet回退'),
]):
    ac(ws2, r+i, 2, idx, font=FONT_INDEX, fill=FILL_SECTION, alignment=ALIGN_CENTER)
    ws2.merge_cells(start_row=r+i, start_column=3, end_row=r+i, end_column=MAIN_COLS)
    ac(ws2, r+i, 3, content, font=FONT_NORMAL)
r += 2

write_header_row(ws2, r, HDR)
r += 1

cube_rows = [
    ('部署阶段', 1, '解压CubeStudio源码', M1, 'root',
     '''cd /data/offline
unzip -o cube-studio-master.zip -d /root/''',
     'ls /root/cube-studio-master/install/kubernetes/'),

    ('部署阶段', 2, 'CubeStudio配置调整', M1, 'root',
     '''# 修改 /root/cube-studio-master/install/kubernetes/cube/overlays/config/config.py
K8S_NETWORK_MODE = "ipvs"
CONTAINER_CLI = "nerdctl"''',
     '''grep -E "K8S_NETWORK_MODE|CONTAINER_CLI" \\
  /root/cube-studio-master/install/kubernetes/cube/overlays/config/config.py'''),

    ('部署阶段', 3, '执行CubeStudio部署', M1, 'root',
     '''cp /etc/kubernetes/admin.conf /root/cube-studio-master/install/kubernetes/config
cd /root/cube-studio-master/install/kubernetes
sh start.sh 10.240.9.13''',
     '''平台入口为istio-system命名空间中的istio-ingressgateway
kubectl get pods -n cube-studio
kubectl get svc -n istio-system istio-ingressgateway'''),

    ('验证阶段', 1, '集群状态检查', M1, 'root',
     '''kubectl get nodes                    # 5节点全部Ready
kubectl get pods -n kube-system      # 全部Running
kubectl get pods -n cube-studio      # 全部Running''',
     '''5个节点Status均为Ready
kube-system和cube-studio下Pod均为Running'''),

    ('验证阶段', 2, '前端访问验证', M1, 'root',
     '''curl http://10.240.9.13:30080''',
     '''浏览器访问 http://10.240.9.13:30080
返回CubeStudio登录页 默认账号 admin / admin'''),

    ('回退阶段', 1, '卸载CubeStudio', M1, 'root',
     '''kubectl delete namespace cube-studio
rm -rf /data/k8s-volumes/{mysql,redis,cubestudio}/*''',
     'kubectl get ns cube-studio  # 已删除'),
]

r = write_phase_steps(ws2, r, cube_rows)
freeze_and_print(ws2, 'A8')

# ========================================
# Sheets 3-9: same structure as before
# ========================================
ws3 = wb.create_sheet('服务器与软件版本')
set_col_widths(ws3, [16, 16, 22, 12, 16, 22])
r = 1
merge_title(ws3, r, 6, '服务器规划')
r += 1
write_header_row(ws3, r, ['IP地址', '主机名', '角色', 'CPU/内存', '操作系统', '备注'])
r += 1
for row_data in [
    ('10.240.9.13', 'k8s-master1', 'Master + Harbor', '40C/256G', 'CentOS 7.7.1908', 'Harbor部署于此节点'),
    ('10.240.9.14', 'k8s-master2', 'Master', '40C/256G', 'CentOS 7.7.1908', ''),
    ('10.240.9.15', 'k8s-master3', 'Master', '40C/256G', 'CentOS 7.7.1908', ''),
    ('10.240.9.16', 'k8s-worker1', 'Worker', '40C/256G', 'CentOS 7.7.1908', ''),
    ('10.240.9.49', 'k8s-worker2', 'Worker', '40C/256G', 'CentOS 7.7.1908', ''),
]:
    for c, v in enumerate(row_data):
        ac(ws3, r, c+1, v, alignment=ALIGN_CENTER)
    set_row_h(ws3, r, 22)
    r += 1
r += 1
merge_title(ws3, r, 6, '软件版本')
r += 1
write_header_row(ws3, r, ['组件', '版本', '说明', '', '', ''])
r += 1
for comp, ver, desc in [
    ('Kubernetes', '1.28.2', 'kubeadm部署，3Master+2Worker高可用'),
    ('Containerd', '1.7.29', '容器运行时'),
    ('CNI Plugins', '1.1.1', '网络插件基础组件'),
    ('Flannel', 'v0.28.4', 'CNI网络插件'),
    ('Harbor', '2.11.1', '集群内镜像仓库（部署于Master-1）'),
    ('CubeStudio Backend', '2026.03.01', 'CubeStudio后端服务'),
    ('CubeStudio Frontend', '2026.03.01', 'CubeStudio前端服务'),
    ('MySQL', '8.0.32', '容器化部署，后续独立部署'),
    ('Redis', '7.4', '容器化部署，后续独立部署'),
]:
    ac(ws3, r, 1, comp, alignment=ALIGN_CENTER)
    ac(ws3, r, 2, ver, alignment=ALIGN_CENTER)
    ws3.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    ac(ws3, r, 3, desc)
    set_row_h(ws3, r, 22)
    r += 1
freeze_and_print(ws3, 'A2')

# Sheet 4: 物料清单
ws4 = wb.create_sheet('物料清单')
set_col_widths(ws4, [48, 38, 12, 12, 25])
r = 1
merge_title(ws4, r, 5, '物料清单 — /data/offline/')
r += 1
write_header_row(ws4, r, ['目录/文件', '说明', '节点数', '大小(约)', '分发方式'])
r += 1
for path, desc, count, size, method in [
    ('nerdctl-full-1.7.6-linux-amd64.tar.gz', 'Nerdctl完整包（含containerd/buildkit等）', '5', '80MB', 'scp分发'),
    ('k8s-bin/', 'kubeadm(45MB)、kubelet(110MB)、kubectl(45MB)', '5', '200MB', 'scp分发'),
    ('cni/', 'cni-plugins-linux-amd64-v1.1.1.tgz', '5', '38MB', 'scp分发'),
    ('containerd/', 'containerd-1.7.29 + runc.amd64', '5', '60MB', 'scp分发'),
    ('harbor/', 'harbor-offline-installer-v2.11.1.tgz', '1（Master-1）', '-', 'scp分发'),
    ('images/', '容器镜像.tar文件（54个）', '5', '-', 'scp分发'),
    ('model/', 'coco.zip、resnet50等模型文件', '5', '-', 'scp分发'),
    ('cube-studio-master.zip', 'CubeStudio源码', '1（Master-1）', '200MB', 'scp分发'),
    ('kube-flannel.yml', 'Flannel部署YAML文件', '1（Master-1）', '15KB', 'scp分发'),
    ('nfsrpm.tar.gz', 'NFS相关RPM包', '5', '-', 'scp分发'),
    ('rpm-new-centos/', 'CentOS 7.x离线RPM包（含Docker CE等，约600+个）', '1（Master-1）', '-', 'scp分发'),
]:
    ac(ws4, r, 1, path, font=FONT_MONO)
    ac(ws4, r, 2, desc)
    ac(ws4, r, 3, count, alignment=ALIGN_CENTER)
    ac(ws4, r, 4, size, alignment=ALIGN_CENTER)
    ac(ws4, r, 5, method, alignment=ALIGN_CENTER)
    set_row_h(ws4, r, 22)
    r += 1
freeze_and_print(ws4, 'A3')

# Sheet 5: 镜像列表
ws5 = wb.create_sheet('镜像列表')
set_col_widths(ws5, [8, 48, 30, 22, 40])
r = 1
merge_title(ws5, r, 5, '容器镜像列表 — 共54个镜像（离线 .tar 文件）')
r += 1
write_header_row(ws5, r, ['序号', '镜像名', 'Tag', '分类', '说明/文件名'])
r += 1
images_data = [
    (1, 'registry.cn-hangzhou.aliyuncs.com/google_containers/kube-apiserver', 'v1.28.2', 'K8s核心', 'K8s API Server'),
    (2, 'registry.cn-hangzhou.aliyuncs.com/google_containers/kube-controller-manager', 'v1.28.2', 'K8s核心', 'Controller Manager'),
    (3, 'registry.cn-hangzhou.aliyuncs.com/google_containers/kube-scheduler', 'v1.28.2', 'K8s核心', 'Scheduler'),
    (4, 'registry.cn-hangzhou.aliyuncs.com/google_containers/kube-proxy', 'v1.28.2', 'K8s核心', 'Proxy'),
    (5, 'registry.cn-hangzhou.aliyuncs.com/google_containers/coredns', 'v1.10.1', 'K8s核心', 'CoreDNS'),
    (6, 'registry.cn-hangzhou.aliyuncs.com/google_containers/etcd', '3.5.9-0', 'K8s核心', 'etcd'),
    (7, 'registry.cn-hangzhou.aliyuncs.com/google_containers/pause', '3.9', 'K8s核心', 'Pause容器'),
    (8, 'registry.cn-hangzhou.aliyuncs.com/google_containers/pause', '3.8', 'K8s核心', 'Pause备用'),
    (9, 'ghcr.io/flannel-io/flannel', 'v0.28.4', 'CNI网络', 'Flannel主程序'),
    (10, 'ghcr.io/flannel-io/flannel-cni-plugin', 'v1.9.1-flannel1', 'CNI网络', 'Flannel CNI插件'),
    (11, 'ccr.ccs.tencentyun.com/cube-studio/kubeflow-dashboard', '2026.03.01', 'CubeStudio核心', '后端服务'),
    (12, 'ccr.ccs.tencentyun.com/cube-studio/kubeflow-dashboard-frontend', '2026.03.01', 'CubeStudio核心', '前端服务'),
    (13, 'ccr.ccs.tencentyun.com/cube-studio/kubeflow-dashboard', '2025.03.01', 'CubeStudio核心', '后端（旧版）'),
    (14, 'ccr.ccs.tencentyun.com/cube-studio/kubeflow-dashboard-frontend', '2025.03.01', 'CubeStudio核心', '前端（旧版）'),
    (15, 'docker.io/library/mysql', '8.0.32', '基础服务', 'MySQL数据库'),
    (16, 'ccr.ccs.tencentyun.com/cube-studio/redis', '7.4', '基础服务', 'Redis缓存'),
    (17, 'docker.io/library/busybox', '1.36.0', '基础服务', 'Busybox工具'),
    (18, 'docker.io/library/alpine', '3.10', '基础服务', 'Alpine基础镜像'),
    (19, 'docker.io/library/python', '3.9', '基础服务', 'Python基础镜像'),
    (20, 'docker.io/library/ubuntu', '20.04', '基础服务', 'Ubuntu基础镜像'),
    (21, 'docker.io/library/docker', '23.0.4', '基础服务', 'Docker-dind镜像'),
    (22, 'docker.io/library/nginx', 'latest', '基础服务', 'Nginx镜像'),
    (23, 'docker.io/library/postgres', '11.5', '基础服务', 'PostgreSQL数据库'),
    (24, 'docker.io/grafana/grafana', '9.5.20', '监控', 'Grafana监控面板'),
    (25, 'docker.io/prom/prometheus', 'v2.27.1', '监控', 'Prometheus监控'),
    (26, 'docker.io/prom/node-exporter', 'v1.5.0', '监控', 'Node Exporter'),
    (27, 'quay.io/prometheus-operator/prometheus-operator', 'v0.46.0', '监控', 'Prometheus Operator'),
    (28, 'quay.io/prometheus-operator/prometheus-config-reloader', 'v0.46.0', '监控', 'Prometheus配置重载'),
    (29, 'docker.io/minio/minio', 'RELEASE.2023-04-20T17-56-55Z', '存储', 'MinIO对象存储'),
    (30, 'docker.io/kubernetesui/dashboard', 'v2.6.1', 'Dashboard', 'K8s Dashboard'),
    (31, 'docker.io/kubernetesui/metrics-scraper', 'v1.0.8', 'Dashboard', 'Dashboard Metrics'),
    (32, 'ccr.ccs.tencentyun.com/cube-studio/k8s-dashboard', 'v2.6.1', 'Dashboard', 'CubeStudio定制Dashboard'),
    (33, 'ccr.ccs.tencentyun.com/cube-studio/kube-rbac-proxy', '0.14.1', 'Dashboard', 'RBAC代理'),
    (34, 'docker.io/istio/pilot', '1.15.0', 'Istio', 'Istio Pilot'),
    (35, 'docker.io/istio/proxyv2', '1.15.0', 'Istio', 'Istio Proxy'),
    (36, 'docker.io/kubeflow/training-operator', 'v1-8a066f9', '训练', '训练Operator'),
    (37, 'ccr.ccs.tencentyun.com/cube-studio/nni', '20240501', '训练', 'NNI调参工具'),
    (38, 'docker.io/volcanosh/vc-controller-manager', 'v1.7.0', '调度', 'Volcano控制器'),
    (39, 'docker.io/volcanosh/vc-scheduler', 'v1.7.0', '调度', 'Volcano调度器'),
    (40, 'docker.io/volcanosh/vc-webhook-manager', 'v1.7.0', '调度', 'Volcano Webhook'),
    (41, 'nvcr.io/nvidia/dcgm-exporter', '3.1.7-3.1.4-ubuntu20.04', 'GPU', 'GPU监控'),
    (42, 'nvcr.io/nvidia/k8s-device-plugin', 'v0.11.0-ubuntu20.04', 'GPU', 'GPU设备插件'),
    (43, 'docker.io/carlosedp/addon-resizer', 'v1.8.4', '其他', '资源调整器'),
    (44, 'ccr.ccs.tencentyun.com/cube-argoproj/workflow-controller', 'v3.4.3', '工作流', 'Argo Workflow Controller'),
    (45, 'ccr.ccs.tencentyun.com/cube-argoproj/argoexec', 'v3.4.3', '工作流', 'Argo Executor'),
    (46, 'ccr.ccs.tencentyun.com/cube-argoproj/argocli', 'v3.4.3', '工作流', 'Argo CLI'),
    (47, 'ccr.ccs.tencentyun.com/cube-studio/torchserve', '0.7.1-cpu', '推理服务', 'TorchServe推理'),
    (48, 'ccr.ccs.tencentyun.com/cube-studio/tritonserver', '22.07-py3', '推理服务', 'Triton推理服务'),
    (49, 'ccr.ccs.tencentyun.com/cube-studio/notebook-jupyter-ubuntu-cpu', '1.0.0', 'Notebook', 'Jupyter Notebook CPU'),
    (50, 'ccr.ccs.tencentyun.com/cube-studio/notebook-jupyter-ubuntu22.04', 'latest', 'Notebook', 'Jupyter Notebook Ubuntu22'),
    (51, 'ccr.ccs.tencentyun.com/cube-studio/notebook-jupyter-ubuntu-bigdata', 'latest', 'Notebook', 'Jupyter Notebook 大数据'),
    (52, 'ccr.ccs.tencentyun.com/cube-studio/notebook-jupyter-ubuntu22.04-cuda11.8.0-cudnn8', 'latest', 'Notebook', 'Jupyter GPU'),
    (53, 'ccr.ccs.tencentyun.com/cube-studio/notebook-vscode-ubuntu-cpu-base', 'latest', 'Notebook', 'VSCode Notebook CPU'),
    (54, 'ccr.ccs.tencentyun.com/cube-studio/notebook-vscode-ubuntu-gpu-base', 'latest', 'Notebook', 'VSCode Notebook GPU'),
]
for ridx, (idx, img, tag, cat, desc) in enumerate(images_data):
    fill_row = FILL_ALT_ROW if ridx % 2 == 0 else None
    ac(ws5, r, 1, idx, font=FONT_NORMAL, fill=fill_row, alignment=ALIGN_CENTER)
    ac(ws5, r, 2, img, font=FONT_MONO, fill=fill_row)
    ac(ws5, r, 3, tag, font=FONT_MONO, fill=fill_row, alignment=ALIGN_CENTER)
    ac(ws5, r, 4, cat, font=FONT_INDEX, fill=fill_row, alignment=ALIGN_CENTER)
    ac(ws5, r, 5, desc, font=FONT_NORMAL, fill=fill_row)
    set_row_h(ws5, r, 18)
    r += 1
r += 1
ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
ac(ws5, r, 1, '汇总：K8s核心 8个 | CNI网络 2个 | CubeStudio核心 4个 | 基础服务 9个 | 监控 5个 | 存储 1个 | Dashboard 4个 | Istio 2个 | 训练 2个 | 调度(Volcano) 3个 | GPU 2个 | 工作流(Argo) 3个 | 推理 2个 | Notebook 6个 | 其他 1个 = 54个',
    font=FONT_SECTION, fill=FILL_SECTION, alignment=ALIGN_LEFT)
set_row_h(ws5, r, 25)
freeze_and_print(ws5, 'A3')

# Sheet 6: RPM包及yum依赖
ws6 = wb.create_sheet('RPM包及yum依赖')
set_col_widths(ws6, [18, 45, 20, 55])
r = 1
merge_title(ws6, r, 4, 'RPM包及yum依赖 — 离线部署所需系统包')
r += 1
ws6.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
ac(ws6, r, 1, '一、yum在线安装依赖包（yum源可用时执行，所有节点）', font=FONT_SECTION, fill=FILL_SECTION, alignment=ALIGN_LEFT)
set_row_h(ws6, r, 25)
r += 1
write_header_row(ws6, r, ['包名', '用途', '安装命令', '说明'])
r += 1
for ridx, (pkg, purpose, cmd, note) in enumerate([
    ('yum-utils', 'YUM工具集', 'yum install -y yum-utils', 'yum-config-manager工具'),
    ('device-mapper-persistent-data', '存储设备映射', 'yum install -y device-mapper-persistent-data', 'LVM2依赖'),
    ('lvm2', '逻辑卷管理', 'yum install -y lvm2', '存储管理'),
    ('iptables', '防火墙规则管理', 'yum install -y iptables', 'kube-proxy依赖'),
    ('container-selinux', '容器SELinux策略', 'yum install -y container-selinux', '容器运行时SELinux策略'),
    ('iptables-services', 'iptables服务管理', 'yum install -y iptables-services', 'iptables服务'),
    ('socat', '网络工具', 'yum install -y socat', 'kubeadm init依赖'),
    ('conntrack-tools', '连接跟踪工具', 'yum install -y conntrack-tools', 'kube-proxy依赖'),
    ('ipvsadm', 'IPVS管理工具', 'yum install -y ipvsadm', 'IPVS模式K8s需要'),
    ('ipset', 'IP集合管理', 'yum install -y ipset', 'kube-proxy ipvs依赖'),
    ('libseccomp', '安全计算库', 'yum install -y libseccomp', '容器安全策略'),
    ('ebtables', '以太网桥防火墙', 'yum install -y ebtables', 'kube-proxy依赖'),
    ('ethtool', '网卡配置工具', 'yum install -y ethtool', '网络管理'),
]):
    fill_row = FILL_ALT_ROW if ridx % 2 == 0 else None
    ac(ws6, r, 1, pkg, font=FONT_MONO, fill=fill_row)
    ac(ws6, r, 2, purpose, font=FONT_NORMAL, fill=fill_row)
    ac(ws6, r, 3, cmd, font=FONT_MONO, fill=fill_row)
    ac(ws6, r, 4, note, font=FONT_NORMAL, fill=fill_row)
    set_row_h(ws6, r, 22)
    r += 1
r += 1
ws6.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
ac(ws6, r, 1, '一键安装命令：yum install -y yum-utils device-mapper-persistent-data lvm2 iptables container-selinux iptables-services socat conntrack-tools ipvsadm ipset libseccomp ebtables ethtool',
    font=Font(name='Consolas', size=11, bold=True, color='FFC00000'), fill=FILL_SECTION, alignment=ALIGN_LEFT)
set_row_h(ws6, r, 28)
r += 2
ws6.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
ac(ws6, r, 1, '二、离线RPM包（无yum源时使用 — /data/offline/）', font=FONT_SECTION, fill=FILL_SECTION, alignment=ALIGN_LEFT)
set_row_h(ws6, r, 25)
r += 1
write_header_row(ws6, r, ['目录', '适用系统', '数量', '说明'])
r += 1
for ridx, (dir_name, system, count, note) in enumerate([
    ('rpm-new-centos/', 'CentOS 7.x', '约600+个.rpm', 'CentOS 7.x离线RPM包，含repodata/元数据目录。包含Docker CE、iptables、conntrack-tools、socat等全套系统组件。'),
    ('rpm-new-kilyn/', '麒麟V10', '23个.rpm', '麒麟V10系统RPM包（备用）'),
    ('rpm-old-kilyn/', '麒麟V10（旧版）', '17个.rpm', '麒麟V10系统RPM包旧版本（兼容性备用）'),
]):
    fill_row = FILL_ALT_ROW if ridx % 2 == 0 else None
    ac(ws6, r, 1, dir_name, font=FONT_MONO, fill=fill_row)
    ac(ws6, r, 2, system, font=FONT_NORMAL, fill=fill_row, alignment=ALIGN_CENTER)
    ac(ws6, r, 3, count, font=FONT_NORMAL, fill=fill_row, alignment=ALIGN_CENTER)
    ac(ws6, r, 4, note, font=FONT_NORMAL, fill=fill_row)
    set_row_h(ws6, r, 35)
    r += 1
r += 1
ws6.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
ac(ws6, r, 1, '三、离线安装命令', font=FONT_SECTION, fill=FILL_SECTION, alignment=ALIGN_LEFT)
set_row_h(ws6, r, 25)
r += 1
ws6.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
ac(ws6, r, 1, '# CentOS 7.x — 安装Docker RPM（Master-1，Harbor依赖）\ncd /data/offline/rpm-new-centos\nrpm -ivh docker-ce-*.rpm docker-ce-cli-*.rpm containerd.io-*.rpm docker-compose-plugin-*.rpm --nodeps\n\n# 有yum源时可替代为：\n# yum install -y docker-ce docker-ce-cli containerd.io docker-compose-plugin',
    font=FONT_MONO, alignment=ALIGN_LEFT_TOP)
set_row_h(ws6, r, 100)
freeze_and_print(ws6, 'A3')

# Sheet 7: 常用命令
ws7 = wb.create_sheet('常用命令')
set_col_widths(ws7, [25, 65, 25])
r = 1
merge_title(ws7, r, 3, '常用命令参考')
r += 1
write_header_row(ws7, r, ['类别', '命令', '说明'])
r += 1
for ridx, (cat, cmd, desc) in enumerate([
    ('集群管理', 'kubectl get nodes -o wide', '查看所有节点详细信息'),
    ('集群管理', 'kubectl get pods -A -o wide', '查看所有命名空间的Pod'),
    ('集群管理', 'kubectl describe node <node>', '查看节点详细信息'),
    ('日志查看', 'kubectl logs <pod> -n <ns> --tail=100', '查看Pod最近100行日志'),
    ('日志查看', 'journalctl -u kubelet -f', '实时查看kubelet日志'),
    ('日志查看', 'journalctl -u containerd -f', '实时查看containerd日志'),
    ('镜像管理', 'ctr -n k8s.io images ls', '列出containerd中所有镜像'),
    ('镜像管理', 'ctr -n k8s.io images import <file>.tar', '导入镜像到containerd'),
    ('Token管理', 'kubeadm token create --print-join-command', '生成新的join命令'),
    ('Harbor', 'docker login 10.240.9.13 -u admin -p Harbor12345', '登录Harbor'),
    ('系统检查', 'getenforce', '查看SELinux状态'),
    ('系统检查', 'systemctl status firewalld', '查看防火墙状态'),
    ('系统检查', 'showmount -e 10.240.9.13', '查看NFS共享'),
    ('故障排查', 'kubectl describe pod <pod> -n <ns>', '排查Pod启动异常'),
    ('故障排查', 'kubeadm reset -f', '重置K8s节点'),
]):
    fill_row = FILL_ALT_ROW if ridx % 2 == 0 else None
    ac(ws7, r, 1, cat, font=FONT_NORMAL, fill=fill_row, alignment=ALIGN_CENTER)
    ac(ws7, r, 2, cmd, font=FONT_MONO, fill=fill_row)
    ac(ws7, r, 3, desc, font=FONT_NORMAL, fill=fill_row)
    set_row_h(ws7, r, 22)
    r += 1
freeze_and_print(ws7, 'A3')

# Sheet 8: 常见问题处理
ws8 = wb.create_sheet('常见问题处理')
set_col_widths(ws8, [22, 50, 50, 25])
r = 1
merge_title(ws8, r, 4, '常见问题处理')
r += 1
write_header_row(ws8, r, ['问题', '现象/原因', '处理方法', '备注'])
r += 1
for ridx, (problem, cause, solution, note) in enumerate([
    ('节点NotReady', 'kubelet异常、CNI未安装、镜像未导入、containerd异常',
     '''kubectl describe node <node-name>
journalctl -u kubelet -f
# 1. CNI未安装 → 检查/opt/cni/bin/
# 2. 镜像未导入 → ctr -n k8s.io images ls
# 3. containerd异常 → systemctl restart containerd
systemctl restart kubelet''', '最常见节点问题'),
    ('镜像拉取失败', 'Pod处于ImagePullBackOff/ErrImagePull',
     '''kubectl describe pod <pod> -n <ns> | grep -i image
ctr -n k8s.io images ls | grep <image-name>
# 缺镜像：cd /data/offline/images && ctr -n k8s.io images import <file>.tar
# 检查Harbor：cat /etc/containerd/certs.d/10.240.9.13/hosts.toml''', '确保54个镜像全部导入'),
    ('Token过期', 'kubeadm join报token无效',
     '''kubeadm token create --print-join-command''', 'Token 24小时内有效'),
    ('Harbor无法访问', 'Harbor容器未启动或端口不通',
     '''docker ps | grep harbor
cd /opt/harbor
docker-compose down && docker-compose up -d
docker-compose logs --tail=50''', 'Harbor在Master-1'),
    ('CNI网络异常', 'Pod间无法通信',
     '''kubectl get pods -n kube-system | grep flannel
kubectl logs <flannel-pod> -n kube-system
kubectl delete pod <flannel-pod> -n kube-system''', '确保Flannel已部署'),
    ('Containerd启动失败', 'systemctl status containerd failed',
     '''containerd config dump > /dev/null
journalctl -u containerd -n 50
# 常见：config.toml语法错误 / CNI缺失 / runc未安装''', '修改config.toml后重启'),
]):
    fill_row = FILL_ALT_ROW if ridx % 2 == 0 else None
    ac(ws8, r, 1, problem, font=FONT_INDEX, fill=fill_row, alignment=ALIGN_CENTER)
    ac(ws8, r, 2, cause, font=FONT_NORMAL, fill=fill_row, alignment=ALIGN_LEFT_TOP)
    ac(ws8, r, 3, solution, font=FONT_NORMAL, fill=fill_row, alignment=ALIGN_LEFT_TOP)
    ac(ws8, r, 4, note, font=FONT_NORMAL, fill=fill_row, alignment=ALIGN_LEFT_TOP)
    set_row_h(ws8, r, 140)
    r += 1
freeze_and_print(ws8, 'A3')

# Sheet 9: 附录
ws9 = wb.create_sheet('附录')
set_col_widths(ws9, [18, 25, 50, 30])
r = 1
merge_title(ws9, r, 4, '附录 — 关键端口与目录结构')
r += 1
ws9.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
ac(ws9, r, 1, '关键端口', font=FONT_SECTION, fill=FILL_SECTION, alignment=ALIGN_CENTER)
set_row_h(ws9, r, 25)
r += 1
write_header_row(ws9, r, ['端口', '服务', '说明', '访问方式'])
r += 1
for port, svc, desc, access in [
    ('6443', 'K8s API Server', 'Kubernetes API入口', '集群内部'),
    ('2379-2380', 'etcd', 'etcd集群通信', '集群内部'),
    ('30080', 'CubeStudio前端', 'CubeStudio Web界面', '浏览器 http://10.240.9.13:30080'),
    ('80', 'Harbor', 'Harbor镜像仓库', '浏览器 http://10.240.9.13'),
    ('3306', 'MySQL', '数据库（集群内）', '集群内部'),
    ('6379', 'Redis', '缓存（集群内）', '集群内部'),
]:
    ac(ws9, r, 1, port, font=FONT_NORMAL, alignment=ALIGN_CENTER)
    ac(ws9, r, 2, svc, font=FONT_NORMAL, alignment=ALIGN_CENTER)
    ac(ws9, r, 3, desc)
    ac(ws9, r, 4, access)
    set_row_h(ws9, r, 22)
    r += 1
r += 2
ws9.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
ac(ws9, r, 1, '目录结构（/data/）', font=FONT_SECTION, fill=FILL_SECTION, alignment=ALIGN_CENTER)
set_row_h(ws9, r, 25)
r += 1
write_header_row(ws9, r, ['目录', '用途', '内容', '所在节点'])
r += 1
for ridx, (path, purpose, content, node) in enumerate([
    ('/data/offline/', '离线物料根目录', 'k8s-bin/ cni/ containerd/ harbor/ images/ model/', '所有节点'),
    ('/data/offline/k8s-bin/', 'K8s二进制', 'kubeadm kubelet kubectl', '所有节点'),
    ('/data/offline/images/', '容器镜像', '54个.tar镜像文件', '所有节点'),
    ('/data/offline/harbor/', 'Harbor安装包', 'harbor-offline-installer-v2.11.1.tgz', 'Master-1'),
    ('/data/k8s-volumes/', '持久化存储', 'mysql/ redis/ cubestudio/', '所有节点'),
    ('/data/nfs/', 'NFS共享存储', 'k8s/ → /data/nfs/k8s/', '所有节点'),
    ('/opt/harbor/', 'Harbor安装目录', 'Harbor配置文件及数据', 'Master-1'),
    ('/root/cube-studio-master/', 'CubeStudio源码', 'install/kubernetes/cube/', 'Master-1'),
]):
    fill_row = FILL_ALT_ROW if ridx % 2 == 0 else None
    ac(ws9, r, 1, path, font=FONT_MONO, fill=fill_row)
    ac(ws9, r, 2, purpose, font=FONT_NORMAL, fill=fill_row)
    ac(ws9, r, 3, content, font=FONT_NORMAL, fill=fill_row)
    ac(ws9, r, 4, node, font=FONT_NORMAL, fill=fill_row, alignment=ALIGN_CENTER)
    set_row_h(ws9, r, 22)
    r += 1
freeze_and_print(ws9, 'A3')

# ==================== SAVE ====================
output_path = 'c:/Users/27404/Desktop/BOS/cubeStudio/cube-studio-master/CubeStudio离线部署操作手册-SIT环境-v1.1.xlsx'
wb.save(output_path)
print(f'文件已生成: {output_path}')
print(f'包含 {len(wb.sheetnames)} 个Sheet:')
for i, name in enumerate(wb.sheetnames):
    print(f'  {i+1}. {name}')
